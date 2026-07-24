import requests
import time
import json
import sqlite3
import os
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from flask import Flask, request
import threading

# ===== ЗАПУСК FLASK (для Render) =====
app = Flask(__name__)

@app.route('/')
def home():
    return "Бот работает!"

@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.get_json()
    if data and "message" in data:
        msg = data["message"]
        chat_id = msg["chat"]["id"]
        text = msg.get("text", "")
        if text:
            process_text_command(chat_id, text)
    return "ok", 200

def run_flask():
    app.run(host='0.0.0.0', port=10000)

threading.Thread(target=run_flask, daemon=True).start()

# ===== ТОКЕН И АДМИН =====
TG_TOKEN = os.environ.get("TG_TOKEN")
ADMIN_ID = 7461823442
REVIEW_GROUP_ID = -1004397763875

if not TG_TOKEN:
    print("❌ Ошибка: не установлен TG_TOKEN")
    exit()

print("🚀 Запуск финансового помощника...")

# ===== ПРОВЕРКА TELEGRAM =====
url = f"https://api.telegram.org/bot{TG_TOKEN}/getMe"
resp = requests.get(url)
if not resp.json().get("ok"):
    print("❌ Ошибка: неверный Telegram-токен")
    exit()

print("✅ Telegram подключён")

# ===== УСТАНОВКА WEBHOOK =====
WEBHOOK_URL = "https://finance-bot-ldes.onrender.com/webhook"
print(f"🔗 Устанавливаю Webhook: {WEBHOOK_URL}")
webhook_url = f"https://api.telegram.org/bot{TG_TOKEN}/setWebhook?url={WEBHOOK_URL}"
resp = requests.get(webhook_url)
if resp.json().get("ok"):
    print(f"✅ Webhook установлен: {WEBHOOK_URL}")
else:
    print("❌ Ошибка установки Webhook:", resp.json())

# ===== БАЗА ДАННЫХ =====
DB_PATH = "finance.db"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY,
        premium INTEGER DEFAULT 0,
        premium_until TEXT,
        first_visit TEXT,
        last_activity TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        type TEXT,
        category TEXT,
        amount REAL,
        description TEXT,
        date TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS budgets (
        user_id INTEGER,
        category TEXT,
        limit_amount REAL,
        PRIMARY KEY (user_id, category)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS donations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        amount INTEGER,
        date TEXT
    )''')
    conn.commit()
    conn.close()

init_db()

# ===== ПЕРЕМЕННАЯ ТЕХОБСЛУЖИВАНИЯ =====
maintenance_mode = False

# ===== КАТЕГОРИИ =====
CATEGORIES = {
    "income": {
        "💰 Зарплата": "основная работа",
        "💼 Фриланс": "проект, заказ",
        "🎁 Подарок": "деньги на день рождения",
        "📈 Инвестиции": "дивиденды",
        "🏦 Проценты": "банковские проценты",
        "🔄 Возврат": "вернули долг",
        "💰 Другое": "прочий доход"
    },
    "expense": {
        "🍔 Еда": "продукты, кафе",
        "🚇 Транспорт": "такси, бензин",
        "🏠 ЖКХ": "квартплата, свет",
        "🛍️ Покупки": "техника, подарки",
        "💊 Здоровье": "лекарства, врачи",
        "🎉 Развлечения": "кино, игры",
        "📱 Связь": "интернет, телефон",
        "👕 Одежда": "обувь, вещи",
        "📚 Образование": "курсы, книги",
        "💰 Другое": "прочие расходы"
    }
}

# ===== ФУНКЦИЯ ФОРМАТИРОВАНИЯ ЧИСЕЛ =====
def format_amount(amount):
    return f"{amount:,.0f}".replace(",", " ")

# ===== ФУНКЦИЯ ИЗВЛЕЧЕНИЯ ЧИСЛА ИЗ СТРОКИ =====
def extract_amount_and_desc(text):
    match = re.search(r'([\d,\.]+)', text)
    if not match:
        return None, None
    num_str = match.group(1).replace(',', '.')
    try:
        amount = float(num_str)
    except ValueError:
        return None, None
    desc = text[match.end():].strip()
    if desc and desc[0] in ',. ':
        desc = desc[1:].strip()
    return amount, desc

# ===== ФУНКЦИИ БАЗЫ =====
def get_user(user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT premium, premium_until FROM users WHERE user_id = ?', (user_id,))
    result = c.fetchone()
    conn.close()
    return {"premium": result[0], "premium_until": result[1]} if result else None

def create_user(user_id):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('INSERT OR IGNORE INTO users (user_id, premium, premium_until, first_visit, last_activity) VALUES (?, 0, NULL, ?, ?)',
              (user_id, now, now))
    conn.commit()
    conn.close()

def update_activity(user_id):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('UPDATE users SET last_activity = ? WHERE user_id = ?', (now, user_id))
    conn.commit()
    conn.close()

def add_transaction(user_id, trans_type, category, amount, description=""):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute('INSERT INTO transactions (user_id, type, category, amount, description, date) VALUES (?, ?, ?, ?, ?, ?)',
              (user_id, trans_type, category, amount, description, date))
    conn.commit()
    conn.close()
    update_activity(user_id)

def get_balance(user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT SUM(amount) FROM transactions WHERE user_id = ? AND type = "income"', (user_id,))
    income = c.fetchone()[0] or 0
    c.execute('SELECT SUM(amount) FROM transactions WHERE user_id = ? AND type = "expense"', (user_id,))
    expense = c.fetchone()[0] or 0
    conn.close()
    return income - expense, income, expense

def get_monthly_expense(user_id, category):
    month_start = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT SUM(amount) FROM transactions WHERE user_id = ? AND type = "expense" AND category = ? AND date >= ?',
              (user_id, category, month_start))
    result = c.fetchone()[0] or 0
    conn.close()
    return result

def get_transactions_page(user_id, page, per_page=10):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    offset = page * per_page
    c.execute('''SELECT type, category, amount, description, date 
                 FROM transactions WHERE user_id = ? 
                 ORDER BY date DESC LIMIT ? OFFSET ?''',
              (user_id, per_page, offset))
    result = c.fetchall()
    conn.close()
    return result

def get_total_transactions(user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM transactions WHERE user_id = ?', (user_id,))
    result = c.fetchone()[0]
    conn.close()
    return result

def get_report_for_month(user_id, year, month):
    month_start = f"{year}-{month:02d}-01"
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    month_end = f"{next_year}-{next_month:02d}-01"
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT type, category, SUM(amount) FROM transactions WHERE user_id = ? AND date >= ? AND date < ? GROUP BY type, category',
              (user_id, month_start, month_end))
    result = c.fetchall()
    conn.close()
    return result

def get_available_months(user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT DISTINCT strftime("%Y-%m", date) as month FROM transactions WHERE user_id = ? ORDER BY month DESC',
              (user_id,))
    result = [row[0] for row in c.fetchall()]
    conn.close()
    return result

def get_monthly_income_expense(user_id, year, month):
    month_start = f"{year}-{month:02d}-01"
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    month_end = f"{next_year}-{next_month:02d}-01"
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT type, SUM(amount) FROM transactions WHERE user_id = ? AND date >= ? AND date < ? GROUP BY type',
              (user_id, month_start, month_end))
    result = c.fetchall()
    conn.close()
    
    income = 0
    expense = 0
    for trans_type, amount in result:
        if trans_type == "income":
            income = amount
        else:
            expense = amount
    return income, expense

def get_budget(user_id, category):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT limit_amount FROM budgets WHERE user_id = ? AND category = ?', (user_id, category))
    result = c.fetchone()
    conn.close()
    return result[0] if result else None

def set_budget(user_id, category, amount):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('INSERT OR REPLACE INTO budgets (user_id, category, limit_amount) VALUES (?, ?, ?)',
              (user_id, category, amount))
    conn.commit()
    conn.close()

def delete_budget(user_id, category):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DELETE FROM budgets WHERE user_id = ? AND category = ?', (user_id, category))
    conn.commit()
    conn.close()

def delete_all_data(user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DELETE FROM transactions WHERE user_id = ?', (user_id,))
    c.execute('DELETE FROM budgets WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

def add_donation(user_id, amount):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute('INSERT INTO donations (user_id, amount, date) VALUES (?, ?, ?)', (user_id, amount, date))
    conn.commit()
    conn.close()

def get_total_donations():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT SUM(amount) FROM donations')
    total = c.fetchone()[0] or 0
    c.execute('SELECT COUNT(*) FROM donations')
    count = c.fetchone()[0] or 0
    conn.close()
    return total, count

# ===== ВЫГРУЗКА В EXCEL =====
def export_to_excel(chat_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT date, type, category, amount, description FROM transactions WHERE user_id = ? ORDER BY date DESC', (chat_id,))
    data = c.fetchall()
    conn.close()
    
    if not data:
        send_message(chat_id, "📋 Нет операций для выгрузки.", main_keyboard(chat_id))
        return
    
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    months = {}
    for row in data:
        try:
            date_obj = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
            month_key = date_obj.strftime("%Y-%m")
            if month_key not in months:
                months[month_key] = []
            months[month_key].append(row)
        except:
            continue
    
    if not months:
        send_message(chat_id, "📋 Нет операций для выгрузки.", main_keyboard(chat_id))
        return
    
    for month_key, month_data in months.items():
        month_name = datetime.strptime(month_key, "%Y-%m").strftime("%B %Y")
        ws = wb.create_sheet(title=month_name)
        
        headers = ["Дата", "Тип", "Категория", "Сумма", "Описание"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        row_num = 2
        total_income = 0
        total_expense = 0
        
        for row in month_data:
            try:
                date_str = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S").strftime("%d.%m.%Y %H:%M")
            except:
                date_str = row[0]
            trans_type = "Доход" if row[1] == "income" else "Расход"
            amount = row[3]
            formatted_amount = f"{amount:,.0f} р.".replace(",", " ")
            
            ws.cell(row=row_num, column=1, value=date_str)
            ws.cell(row=row_num, column=2, value=trans_type)
            ws.cell(row=row_num, column=3, value=row[2])
            ws.cell(row=row_num, column=4, value=formatted_amount)
            ws.cell(row=row_num, column=5, value=row[4] or "")
            
            if row[1] == "income":
                total_income += amount
            else:
                total_expense += amount
            
            row_num += 1
        
        bold_font = Font(bold=True)
        row_num += 1
        
        ws.cell(row=row_num, column=3, value="ИТОГО ДОХОДЫ:")
        ws.cell(row=row_num, column=4, value=f"{total_income:,.0f} р.".replace(",", " "))
        ws.cell(row=row_num, column=3).font = bold_font
        
        row_num += 1
        ws.cell(row=row_num, column=3, value="ИТОГО РАСХОДЫ:")
        ws.cell(row=row_num, column=4, value=f"{total_expense:,.0f} р.".replace(",", " "))
        ws.cell(row=row_num, column=3).font = bold_font
        
        row_num += 1
        balance = total_income - total_expense
        ws.cell(row=row_num, column=3, value="ДОСТУПНЫЙ БАЛАНС:")
        ws.cell(row=row_num, column=4, value=f"{balance:,.0f} р.".replace(",", " "))
        ws.cell(row=row_num, column=3).font = bold_font
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
    
    filename = f"отчёт_{datetime.now().strftime('%B %Y')}.xlsx"
    wb.save(filename)
    
    url = f"https://api.telegram.org/bot{TG_TOKEN}/sendDocument"
    files = {'document': open(filename, 'rb')}
    data = {'chat_id': chat_id}
    requests.post(url, files=files, data=data)
    
    os.remove(filename)
    send_message(chat_id, "✅ Файл отправлен!", main_keyboard(chat_id))

# ===== СТАТИСТИКА =====
def get_stats():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    c.execute('SELECT COUNT(*) FROM users')
    total_users = c.fetchone()[0]
    
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    c.execute('SELECT COUNT(*) FROM users WHERE last_activity >= ?', (week_ago,))
    active_users = c.fetchone()[0]
    
    month_start = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    c.execute('SELECT COUNT(*) FROM transactions WHERE date >= ?', (month_start,))
    total_ops = c.fetchone()[0]
    
    c.execute('SELECT SUM(amount) FROM transactions WHERE type = "income" AND date >= ?', (month_start,))
    total_income = c.fetchone()[0] or 0
    c.execute('SELECT SUM(amount) FROM transactions WHERE type = "expense" AND date >= ?', (month_start,))
    total_expense = c.fetchone()[0] or 0
    
    conn.close()
    return {
        "total_users": total_users,
        "active_users": active_users,
        "total_ops": total_ops,
        "total_income": total_income,
        "total_expense": total_expense
    }

# ===== КЛАВИАТУРЫ =====
def main_keyboard(chat_id):
    keyboard = [
        ["💰 Баланс", "📊 Отчёт"],
        ["📝 Доход", "💸 Расход"],
        ["📈 Бюджет", "📋 История"],
        ["📤 Выгрузка в Excel", "❓ Инструкция пользователя"],
        ["⭐ Донат", "💬 Отзыв"],
        ["🗑️ Сбросить всё"]
    ]
    if chat_id == ADMIN_ID:
        keyboard.append(["📊 Статистика"])
    return {"keyboard": keyboard, "resize_keyboard": True}

def two_column_keyboard(cats):
    keyboard = []
    for i in range(0, len(cats), 2):
        row = []
        row.append(cats[i])
        if i + 1 < len(cats):
            row.append(cats[i + 1])
        keyboard.append(row)
    keyboard.append(["🔙 Назад"])
    return {"keyboard": keyboard, "resize_keyboard": True}

def category_keyboard(trans_type):
    cats = list(CATEGORIES[trans_type].keys())
    return two_column_keyboard(cats)

def back_keyboard():
    return {"keyboard": [["🔙 Назад"]], "resize_keyboard": True}

def budget_keyboard_with_delete():
    return {"keyboard": [
        ["🗑️ Удалить лимит"],
        ["🔙 Назад"]
    ], "resize_keyboard": True}

def history_keyboard(page, total_pages):
    keyboard = []
    if page > 0:
        keyboard.append(["🔙 Назад"])
    if page < total_pages - 1:
        keyboard.append(["➡️ Дальше"])
    if page > 0 and page < total_pages - 1:
        keyboard = [["🔙 Назад", "➡️ Дальше"]]
    elif page > 0 and page == total_pages - 1:
        keyboard = [["🔙 Назад"]]
    elif page == 0 and page < total_pages - 1:
        keyboard = [["➡️ Дальше"]]
    keyboard.append(["🔙 Главное меню"])
    return {"keyboard": keyboard, "resize_keyboard": True}

def months_keyboard(months):
    keyboard = []
    for month in months:
        year, month_num = month.split("-")
        month_name = datetime(int(year), int(month_num), 1).strftime("%B %Y")
        keyboard.append([f"📆 {month_name}"])
    keyboard.append(["🔙 Назад"])
    return {"keyboard": keyboard, "resize_keyboard": True}

def confirm_delete_keyboard():
    return {"keyboard": [
        ["✅ Да, удалить всё"],
        ["❌ Нет, отмена"]
    ], "resize_keyboard": True}

def send_message(chat_id, text, keyboard=None):
    url = f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage"
    data = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
    if keyboard:
        data["reply_markup"] = json.dumps(keyboard)
    try:
        requests.post(url, data=data, timeout=5)
    except Exception as e:
        print(f"⚠️ Ошибка отправки: {e}")

# ===== ОБРАБОТЧИКИ =====
def process_text_command(chat_id, text):
    """Обрабатывает команды из чата И из Mini App"""
    if not get_user(chat_id):
        create_user(chat_id)

    if maintenance_mode and chat_id != ADMIN_ID:
        send_message(chat_id, "🔧 *Технические работы*\n\nБот временно недоступен.", main_keyboard(chat_id))
        return

    # ===== ДОХОД ИЗ ИНТЕРФЕЙСА =====
    if text.startswith("Доход "):
        parts = text.split(" ", 2)
        if len(parts) >= 2:
            try:
                amount = float(parts[1])
                category = parts[2] if len(parts) > 2 else "💰 Другое"
                desc = ""
                if " " in category:
                    category, desc = category.split(" ", 1)
                add_transaction(chat_id, "income", category, amount, desc)
                send_message(chat_id, f"✅ Доход записан!\n{category}: {format_amount(amount)} р.", main_keyboard(chat_id))
            except:
                send_message(chat_id, "❌ Ошибка формата дохода. Используйте: Доход 15000 Зарплата", main_keyboard(chat_id))
        return

    # ===== РАСХОД ИЗ ИНТЕРФЕЙСА =====
    if text.startswith("Расход "):
        parts = text.split(" ", 2)
        if len(parts) >= 2:
            try:
                amount = float(parts[1])
                category = parts[2] if len(parts) > 2 else "💰 Другое"
                desc = ""
                if " " in category:
                    category, desc = category.split(" ", 1)
                add_transaction(chat_id, "expense", category, amount, desc)
                send_message(chat_id, f"✅ Расход записан!\n{category}: {format_amount(amount)} р.", main_keyboard(chat_id))
            except:
                send_message(chat_id, "❌ Ошибка формата расхода. Используйте: Расход 500 Еда", main_keyboard(chat_id))
        return

    # ===== БЮДЖЕТ ИЗ ИНТЕРФЕЙСА =====
    if text.startswith("Бюджет "):
        parts = text.split(" ", 2)
        if len(parts) >= 3:
            try:
                category = parts[1]
                amount = float(parts[2])
                set_budget(chat_id, category, amount)
                send_message(chat_id, f"✅ Бюджет для {category}: {format_amount(amount)} р.", main_keyboard(chat_id))
            except:
                send_message(chat_id, "❌ Ошибка бюджета. Используйте: Бюджет Еда 10000", main_keyboard(chat_id))
        return

    # ===== ОТЗЫВ ИЗ ИНТЕРФЕЙСА =====
    if text.startswith("Отзыв: "):
        review_text = text[7:]
        user_name = "Пользователь"
        try:
            url = f"https://api.telegram.org/bot{TG_TOKEN}/getChat"
            params = {"chat_id": chat_id}
            response = requests.get(url, params=params)
            user_data = response.json()
            if user_data.get("ok"):
                user_name = user_data.get("result", {}).get("first_name", "Пользователь")
        except:
            pass
        notify_text = (
            f"💬 *Новый отзыв*\n\n"
            f"👤 От: {user_name}\n"
            f"📝 Текст:\n{review_text}"
        )
        send_message(REVIEW_GROUP_ID, notify_text)
        send_message(chat_id, "✅ Спасибо за отзыв! 🙏", main_keyboard(chat_id))
        return

    # ===== ДОНАТ ИЗ ИНТЕРФЕЙСА =====
    if text.startswith("⭐ Донат "):
        try:
            amount = int(text.split(" ")[2])
            add_donation(chat_id, amount)
            send_message(chat_id, f"🙏 Спасибо за донат {amount} Stars!", main_keyboard(chat_id))
        except:
            send_message(chat_id, "❌ Ошибка доната. Используйте: ⭐ Донат 25", main_keyboard(chat_id))
        return

    # ===== СТАТИСТИКА =====
    if text == "📊 Статистика" and chat_id == ADMIN_ID:
        handle_stats(chat_id)
        return

    # ===== ОСТАЛЬНЫЕ КОМАНДЫ (из старого кода) =====
    if text == "/start":
        handle_start(chat_id)
        return
    if text == "/help" or text == "❓ Инструкция пользователя":
        handle_help(chat_id)
        return
    if text == "💰 Баланс":
        handle_balance(chat_id)
        return
    if text == "📊 Отчёт":
        handle_report_current(chat_id)
        return
    if text == "📋 История":
        handle_history(chat_id, 0)
        return
    if text == "📤 Выгрузка в Excel":
        export_to_excel(chat_id)
        return
    if text == "⭐ Донат":
        handle_donate(chat_id)
        return
    if text == "💬 Отзыв":
        handle_review(chat_id)
        return
    if text == "📈 Бюджет":
        handle_budget(chat_id)
        return
    if text == "🗑️ Сбросить всё":
        handle_reset(chat_id)
        return
    if text == "🗑️ Удалить лимит":
        send_message(chat_id, "❌ Удаление лимита доступно только в чате.", main_keyboard(chat_id))
        return
    if text == "✅ Да, удалить всё":
        delete_all_data(chat_id)
        send_message(chat_id, "🗑️ Все данные удалены!", main_keyboard(chat_id))
        return

    # Если ничего не подошло
    send_message(chat_id, "❌ Используйте кнопки меню 👇", main_keyboard(chat_id))

def handle_start(chat_id):
    create_user(chat_id)
    text = (
        "💰 *Финансовый помощник*\n\n"
        "📌 Записывайте доходы и расходы.\n"
        "📊 Стройте отчёты по месяцам.\n"
        "📤 Выгружайте данные в Excel.\n"
        "📈 Устанавливайте лимиты на категории расходов — бот предупредит о превышении расхода.\n\n"
        "Подробная информация — в кнопке «❓ Инструкция пользователя»."
    )
    send_message(chat_id, text, main_keyboard(chat_id))

def handle_help(chat_id):
    text1 = (
        "📖 *Инструкция пользователя (часть 1/3)*\n\n"
        "💰 *Финансовый помощник* помогает вести учёт доходов и расходов.\n\n"
        "🔹 *Как записать доход:*\n"
        "1. Нажмите «📝 Доход»\n"
        "2. Выберите категорию (например, «💰 Зарплата»)\n"
        "3. Введите сумму и описание\n"
        "   *Пример:* 15000 зарплата\n\n"
        "🔹 *Как записать расход:*\n"
        "1. Нажмите «💸 Расход»\n"
        "2. Выберите категорию (например, «🍔 Еда»)\n"
        "3. Введите сумму и описание\n"
        "   *Пример:* 500 обед\n\n"
        "📌 *Если меню не появилось — просто отправьте «Старт» ещё раз.*"
    )
    send_message(chat_id, text1, main_keyboard(chat_id))
    time.sleep(0.5)
    
    text2 = (
        "📖 *Инструкция пользователя (часть 2/3)*\n\n"
        "🔹 *Что такое бюджет?*\n"
        "Это функция установки лимита расходов на категорию в месяц.\n"
        "Нажмите «📈 Бюджет», выберите категорию — бот покажет текущий лимит (если он есть).\n"
        "Вы можете установить новый лимит, изменить его или удалить.\n"
        "Если превысите лимит — бот предупредит.\n\n"
        "🔹 *Как посмотреть отчёт?*\n"
        "Нажмите «📊 Отчёт» — бот покажет текущий месяц.\n"
        "Нажмите «📅 Выбрать месяц» — можно посмотреть любой месяц."
    )
    send_message(chat_id, text2)
    time.sleep(0.5)
    
    text3 = (
        "📖 *Инструкция пользователя (часть 3/3)*\n\n"
        "🔹 *Как посмотреть историю?*\n"
        "Нажмите «📋 История» — бот покажет последние операции.\n"
        "Листайте с помощью кнопок «🔙 Назад» и «➡️ Дальше».\n\n"
        "🔹 *Как удалить все данные?*\n"
        "Нажмите «🗑️ Сбросить всё» — бот запросит подтверждение.\n\n"
        "🔹 *Как выгрузить данные в Excel?*\n"
        "Нажмите «📤 Выгрузка в Excel» — бот пришлёт файл с таблицей.\n"
        "В файле будут все ваши операции: дата, тип, категория, сумма и описание.\n"
        "Внизу таблицы — итоги: общие доходы, расходы и доступный баланс.\n\n"
        "⭐ *Донат:*\n"
        "Вы можете оставить чаевые разработчику — кнопка «⭐ Донат» в меню.\n"
        "💰 Приём донатов осуществляется через Telegram Stars.\n\n"
        "💬 *Отзыв:*\n"
        "Нажмите «💬 Отзыв» и напишите своё мнение о боте.\n"
        "Это поможет сделать его лучше!"
    )
    send_message(chat_id, text3, main_keyboard(chat_id))

def handle_balance(chat_id):
    balance, income, expense = get_balance(chat_id)
    text = (
        f"💰 *Ваш баланс: {format_amount(balance)} р.*\n\n"
        f"↗️ Доходы: {format_amount(income)} р.\n"
        f"↘️ Расходы: {format_amount(expense)} р."
    )
    send_message(chat_id, text, main_keyboard(chat_id))

def handle_report_current(chat_id):
    data = get_report_for_month(chat_id, datetime.now().year, datetime.now().month)
    if not data:
        send_message(chat_id, "📊 За этот месяц операций нет.", main_keyboard(chat_id))
        return
    balance, income, expense = get_balance(chat_id)
    text = f"📊 *Отчёт за {datetime.now().strftime('%B %Y')}*\n\n"
    text += f"↗️ Доходы: {format_amount(income)} р.\n"
    text += f"↘️ Расходы: {format_amount(expense)} р.\n"
    text += f"💵 Свободно: {format_amount(balance)} р.\n\n"
    
    expense_cats = {}
    for trans_type, category, amount in data:
        if trans_type == "expense":
            expense_cats[category] = expense_cats.get(category, 0) + amount
    
    if expense_cats:
        text += "🔹 *Расходы по категориям:*\n"
        for cat, amount in sorted(expense_cats.items(), key=lambda x: x[1], reverse=True):
            budget = get_budget(chat_id, cat)
            line = f"  • {cat}: {format_amount(amount)} р."
            if budget:
                line += f" (лимит: {format_amount(budget)} р.)"
                if amount > budget:
                    line += f" ⚠️ превышен!"
            text += line + "\n"
    else:
        text += "🔹 Расходов нет\n"
    
    send_message(chat_id, text, main_keyboard(chat_id))

def handle_budget(chat_id):
    cats = list(CATEGORIES["expense"].keys())
    send_message(chat_id, "📈 *Выберите категорию для бюджета:*", two_column_keyboard(cats))

def handle_donate(chat_id):
    text = (
        "⭐ *Оставить чаевые*\n\n"
        "Напишите сумму в Stars (например, 10, 25, 50)."
    )
    send_message(chat_id, text, back_keyboard())

def handle_review(chat_id):
    text = (
        "💬 *Оставить отзыв*\n\n"
        "Напишите ваше мнение о боте одним сообщением."
    )
    send_message(chat_id, text, back_keyboard())

def handle_stats(chat_id):
    if chat_id != ADMIN_ID:
        send_message(chat_id, "⛔ Нет доступа.", main_keyboard(chat_id))
        return
    stats = get_stats()
    text = (
        "📊 *Статистика бота*\n\n"
        f"👥 Всего пользователей: {stats['total_users']}\n"
        f"📆 Активных за 7 дней: {stats['active_users']}\n"
        f"📝 Операций за месяц: {stats['total_ops']}\n"
        f"💰 Доходы: {format_amount(stats['total_income'])} р.\n"
        f"📉 Расходы: {format_amount(stats['total_expense'])} р."
    )
    send_message(chat_id, text, main_keyboard(chat_id))

def handle_history(chat_id, page=0):
    total = get_total_transactions(chat_id)
    if total == 0:
        send_message(chat_id, "📋 История пуста.", main_keyboard(chat_id))
        return
    per_page = 10
    total_pages = (total + per_page - 1) // per_page
    if page >= total_pages:
        page = total_pages - 1
    transactions = get_transactions_page(chat_id, page, per_page)
    text = f"📋 *История операций (стр. {page + 1}/{total_pages})*\n\n"
    for trans_type, category, amount, description, date in transactions:
        emoji = "↗️" if trans_type == "income" else "↘️"
        desc = f" ({description})" if description else ""
        text += f"{emoji} {category}: {format_amount(amount)} р.{desc}\n"
    send_message(chat_id, text, history_keyboard(page, total_pages))

def handle_reset(chat_id):
    send_message(
        chat_id,
        "⚠️ *ВНИМАНИЕ!*\n\nВы действительно хотите удалить ВСЕ свои данные?\n"
        "(доходы, расходы, бюджеты)\n\n"
        "Это действие НЕЛЬЗЯ отменить!",
        confirm_delete_keyboard()
    )

def send_invoice(chat_id, amount):
    try:
        url = f"https://api.telegram.org/bot{TG_TOKEN}/sendInvoice"
        payload = {
            "chat_id": chat_id,
            "title": "Поддержка бота",
            "description": "Спасибо за вашу поддержку! ❤️",
            "payload": f"donation_{chat_id}_{amount}",
            "provider_token": "",
            "currency": "XTR",
            "prices": [{"label": f"{amount} Stars", "amount": amount * 100}],
            "start_parameter": "donation"
        }
        response = requests.post(url, json=payload)
        return response.json()
    except Exception as e:
        print(f"⚠️ Ошибка создания инвойса: {e}")
        return None

def handle_pre_checkout_query(pre_checkout_query):
    try:
        url = f"https://api.telegram.org/bot{TG_TOKEN}/answerPreCheckoutQuery"
        payload = {
            "pre_checkout_query_id": pre_checkout_query["id"],
            "ok": True
        }
        requests.post(url, json=payload)
    except Exception as e:
        print(f"⚠️ Ошибка PreCheckout: {e}")

def handle_successful_payment(chat_id, payment_info):
    try:
        amount = payment_info.get("total_amount", 0) // 100
        add_donation(chat_id, amount)
        send_message(chat_id, "🙏 Спасибо за чаевые! ❤️", main_keyboard(chat_id))
    except Exception as e:
        print(f"⚠️ Ошибка обработки платежа: {e}")

print("✅ Бот готов! Ожидаю Webhook-запросы...")
print("=" * 50)
